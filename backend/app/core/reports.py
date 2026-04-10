"""
Rapor üretimi: PDF (WeasyPrint) ve Excel (openpyxl)
"""

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from weasyprint import HTML


_PDF_TEMPLATE = """
<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="utf-8">
<style>
  body { font-family: Arial, sans-serif; font-size: 11px; margin: 20px; color: #222; }
  h1 { font-size: 16px; color: #2c5f2e; border-bottom: 2px solid #2c5f2e; padding-bottom: 6px; }
  h2 { font-size: 13px; color: #2c5f2e; margin-top: 16px; }
  table { width: 100%; border-collapse: collapse; margin-top: 8px; }
  th { background: #2c5f2e; color: white; padding: 5px 8px; text-align: left; }
  td { padding: 4px 8px; border-bottom: 1px solid #e0e0e0; }
  tr:nth-child(even) { background: #f7f7f7; }
  .total { font-weight: bold; background: #e8f5e9; }
  .meta { margin-bottom: 16px; color: #555; }
  .meta span { margin-right: 20px; }
  .warn { color: #c0392b; font-size: 10px; margin-top: 4px; }
  .footer { margin-top: 24px; font-size: 9px; color: #888; border-top: 1px solid #ccc; padding-top: 4px; }
</style>
</head>
<body>
<h1>TMR Rasyon Raporu — {{ ration.name }}</h1>
<div class="meta">
  <span><b>Hayvan:</b> {{ animal.name }}</span>
  <span><b>Tür:</b> {{ "Süt Sığırı" if animal.species == "dairy" else "Besi Sığırı" }}</span>
  <span><b>C.A.:</b> {{ animal.live_weight_kg }} kg</span>
  {% if animal.species == "dairy" %}
  <span><b>Süt:</b> {{ animal.milk_yield_kg_day }} kg/gün</span>
  <span><b>Yağ:</b> {{ animal.fat_pct }}%</span>
  {% else %}
  <span><b>Hedef AGK:</b> {{ animal.target_adg_kg }} kg/gün</span>
  {% endif %}
  <span><b>Tarih:</b> {{ date }}</span>
</div>

<h2>Rasyon Bileşenleri</h2>
<table>
  <tr>
    <th>Hammadde</th>
    <th>Taze (kg)</th>
    <th>KM (kg)</th>
    <th>NEL (Mcal)</th>
    <th>HP (g)</th>
    <th>Ca (g)</th>
    <th>P (g)</th>
    <th>Maliyet (TL)</th>
  </tr>
  {% for item in items %}
  <tr>
    <td>{{ item.name_tr or item.name }}</td>
    <td>{{ "%.2f"|format(item.fresh_kg) }}</td>
    <td>{{ "%.2f"|format(item.dm_kg) }}</td>
    <td>{{ "%.2f"|format(item.nel) if item.nel else "—" }}</td>
    <td>{{ "%.0f"|format(item.cp_g) if item.cp_g else "—" }}</td>
    <td>{{ "%.1f"|format(item.ca_g) if item.ca_g else "—" }}</td>
    <td>{{ "%.1f"|format(item.p_g) if item.p_g else "—" }}</td>
    <td>{{ "%.2f"|format(item.cost) if item.cost else "—" }}</td>
  </tr>
  {% endfor %}
  <tr class="total">
    <td><b>TOPLAM</b></td>
    <td><b>{{ "%.2f"|format(ration.total_fresh_kg) }}</b></td>
    <td><b>{{ "%.2f"|format(ration.total_dm_kg) }}</b></td>
    <td><b>{{ "%.2f"|format(total_nel) }}</b></td>
    <td><b>{{ "%.0f"|format(total_cp) }}</b></td>
    <td><b>{{ "%.1f"|format(total_ca) }}</b></td>
    <td><b>{{ "%.1f"|format(total_p) }}</b></td>
    <td><b>{{ "%.2f"|format(ration.total_cost_tl or 0) }} TL</b></td>
  </tr>
</table>

{% if notes %}
<div class="warn">Not: {{ notes }}</div>
{% endif %}

<div class="footer">
  TMR Rasyon Programı — NRC 2023 |
  Optimizasyon: {{ "LP (Otomatik)" if ration.optimization_mode == "lp" else "Manuel" }} |
  Oluşturulma: {{ date }}
</div>
</body>
</html>
"""


def _render_items(ration):
    rows = []
    total_nel = total_cp = total_ca = total_p = 0.0
    for item in ration.items:
        ing = item.ingredient
        dm_kg = item.dm_weight_kg or (item.fresh_weight_kg * (ing.dm_pct or 100) / 100)
        nel = dm_kg * (ing.nel_mcal_kg or 0)
        cp_g = dm_kg * (ing.cp_pct or 0) * 10
        ca_g = dm_kg * (ing.ca_pct or 0) * 10
        p_g = dm_kg * (ing.p_pct or 0) * 10
        cost = item.fresh_weight_kg * (ing.price_per_kg_tl or 0)
        rows.append({
            "name": ing.name,
            "name_tr": ing.name_tr,
            "fresh_kg": item.fresh_weight_kg,
            "dm_kg": dm_kg,
            "nel": nel if ing.nel_mcal_kg else None,
            "cp_g": cp_g if ing.cp_pct else None,
            "ca_g": ca_g if ing.ca_pct else None,
            "p_g": p_g if ing.p_pct else None,
            "cost": cost if ing.price_per_kg_tl else None,
        })
        total_nel += nel
        total_cp += cp_g
        total_ca += ca_g
        total_p += p_g
    return rows, total_nel, total_cp, total_ca, total_p


def generate_pdf(ration) -> bytes:
    from jinja2 import Template

    rows, total_nel, total_cp, total_ca, total_p = _render_items(ration)
    animal = ration.animal_profile

    html = Template(_PDF_TEMPLATE).render(
        ration=ration,
        animal=animal,
        items=rows,
        total_nel=total_nel,
        total_cp=total_cp,
        total_ca=total_ca,
        total_p=total_p,
        notes=ration.notes,
        date=datetime.now().strftime("%d.%m.%Y"),
    )
    return HTML(string=html).write_pdf()


def generate_excel(ration) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rasyon"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C5F2E")

    headers = ["Hammadde", "Taze (kg)", "KM (kg)", "NEL (Mcal)", "HP (g)", "Ca (g)", "P (g)", "Maliyet (TL)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    rows, total_nel, total_cp, total_ca, total_p = _render_items(ration)
    for row_idx, item in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=item["name_tr"] or item["name"])
        ws.cell(row=row_idx, column=2, value=round(item["fresh_kg"], 3))
        ws.cell(row=row_idx, column=3, value=round(item["dm_kg"], 3))
        ws.cell(row=row_idx, column=4, value=round(item["nel"], 2) if item["nel"] else None)
        ws.cell(row=row_idx, column=5, value=round(item["cp_g"], 0) if item["cp_g"] else None)
        ws.cell(row=row_idx, column=6, value=round(item["ca_g"], 1) if item["ca_g"] else None)
        ws.cell(row=row_idx, column=7, value=round(item["p_g"], 1) if item["p_g"] else None)
        ws.cell(row=row_idx, column=8, value=round(item["cost"], 2) if item["cost"] else None)

    total_row = len(rows) + 2
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="E8F5E9")
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).font = total_font
        ws.cell(row=total_row, column=col).fill = total_fill

    ws.cell(row=total_row, column=1, value="TOPLAM")
    ws.cell(row=total_row, column=2, value=round(ration.total_fresh_kg or 0, 3))
    ws.cell(row=total_row, column=3, value=round(ration.total_dm_kg or 0, 3))
    ws.cell(row=total_row, column=4, value=round(total_nel, 2))
    ws.cell(row=total_row, column=5, value=round(total_cp, 0))
    ws.cell(row=total_row, column=6, value=round(total_ca, 1))
    ws.cell(row=total_row, column=7, value=round(total_p, 1))
    ws.cell(row=total_row, column=8, value=round(ration.total_cost_tl or 0, 2))

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 30
    for col_letter in "BCDEFGH":
        ws.column_dimensions[col_letter].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_mixing_list(ration, herd_size: int = 1) -> dict:
    """Sürü başına günlük ve haftalık karma talimatı."""
    items = []
    for item in ration.items:
        ing = item.ingredient
        daily_kg = item.fresh_weight_kg * herd_size
        items.append({
            "hammadde": ing.name_tr or ing.name,
            "hayvan_basi_kg": round(item.fresh_weight_kg, 2),
            "gunluk_kg": round(daily_kg, 2),
            "haftalik_kg": round(daily_kg * 7, 2),
        })

    return {
        "rasyon_adi": ration.name,
        "hayvan_sayisi": herd_size,
        "toplam_gunluk_kg": round((ration.total_fresh_kg or 0) * herd_size, 2),
        "toplam_haftalik_kg": round((ration.total_fresh_kg or 0) * herd_size * 7, 2),
        "bilesenler": items,
    }
